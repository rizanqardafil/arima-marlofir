"""
Enhanced Export Utilities for BBM Dashboard & MARLOFIR-P
Professional reporting, data export, and visualization utilities
"""

import pandas as pd
import numpy as np
from typing import Dict, List, Tuple, Optional, Any, Union
import logging
from datetime import datetime, timedelta
from pathlib import Path
import json
import base64
from io import BytesIO, StringIO
import zipfile
import tempfile
import warnings
warnings.filterwarnings('ignore')

# Import for advanced exports
try:
    import plotly.graph_objects as go
    import plotly.express as px
    from plotly.subplots import make_subplots
    import plotly.io as pio
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import LineChart, BarChart, Reference
    from openpyxl.drawing.image import Image as ExcelImage
except ImportError as e:
    logging.warning(f"Some export libraries not available: {e}")

# PDF generation (optional)
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics.charts.linecharts import HorizontalLineChart
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False
    logging.warning("ReportLab not available - PDF export disabled")

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class ExportFormat:
    """Supported export formats"""
    EXCEL = "excel"
    PDF = "pdf"
    JSON = "json"
    CSV = "csv"
    HTML = "html"
    PNG = "png"
    SVG = "svg"
    ZIP = "zip"

class ReportTemplate:
    """Report template configurations"""
    EXECUTIVE_SUMMARY = "executive_summary"
    TECHNICAL_DETAILS = "technical_details"
    COMPARISON_REPORT = "comparison_report"
    OPTIMIZATION_ANALYSIS = "optimization_analysis"
    FORECAST_REPORT = "forecast_report"

class ExportUtilities:
    """
    Comprehensive export utilities for BBM Dashboard and MARLOFIR-P
    Handles professional reporting, data export, and visualization
    """
    
    def __init__(self):
        self.temp_dir = Path(tempfile.gettempdir()) / "bbm_exports"
        self.temp_dir.mkdir(exist_ok=True)
        
        # Export configurations
        self.config = {
            'company_name': 'BBM Distribution Analytics',
            'report_author': 'MARLOFIR-P System',
            'logo_path': None,
            'color_scheme': {
                'primary': '#1f77b4',
                'secondary': '#ff7f0e', 
                'success': '#2ca02c',
                'warning': '#d62728',
                'info': '#9467bd'
            },
            'excel_style': {
                'header_fill': 'D3D3D3',
                'header_font': 'Calibri',
                'data_font': 'Calibri',
                'number_format': '#,##0.00'
            }
        }
        
        # Performance tracking
        self.export_stats = {
            'total_exports': 0,
            'successful_exports': 0,
            'failed_exports': 0,
            'formats_used': {},
            'avg_export_time': 0.0
        }
        
        logger.info("Export utilities initialized")
    
    def set_configuration(self, config: Dict[str, Any]) -> None:
        """Update export configuration"""
        self.config.update(config)
        logger.info("Export configuration updated")
    
    def export_arima_results(self, arima_data: Dict[str, Any], 
                           format: str = ExportFormat.EXCEL,
                           template: str = ReportTemplate.FORECAST_REPORT) -> bytes:
        """
        Export ARIMA forecasting results
        
        Args:
            arima_data: ARIMA analysis results
            format: Export format
            template: Report template
            
        Returns:
            Exported data as bytes
        """
        logger.info(f"Exporting ARIMA results in {format} format")
        
        if format == ExportFormat.EXCEL:
            return self._export_arima_excel(arima_data, template)
        elif format == ExportFormat.PDF:
            return self._export_arima_pdf(arima_data, template)
        elif format == ExportFormat.JSON:
            return self._export_arima_json(arima_data)
        elif format == ExportFormat.HTML:
            return self._export_arima_html(arima_data, template)
        else:
            raise ValueError(f"Unsupported format for ARIMA export: {format}")
    
    def export_optimization_results(self, optimization_data: Dict[str, Any],
                                  format: str = ExportFormat.EXCEL,
                                  template: str = ReportTemplate.OPTIMIZATION_ANALYSIS) -> bytes:
        """
        Export optimization results (GA + VRP)
        
        Args:
            optimization_data: Optimization results
            format: Export format
            template: Report template
            
        Returns:
            Exported data as bytes
        """
        logger.info(f"Exporting optimization results in {format} format")
        
        if format == ExportFormat.EXCEL:
            return self._export_optimization_excel(optimization_data, template)
        elif format == ExportFormat.PDF:
            return self._export_optimization_pdf(optimization_data, template)
        elif format == ExportFormat.JSON:
            return self._export_optimization_json(optimization_data)
        else:
            raise ValueError(f"Unsupported format for optimization export: {format}")
    
    def export_combined_report(self, arima_data: Dict[str, Any],
                             optimization_data: Dict[str, Any],
                             format: str = ExportFormat.EXCEL,
                             template: str = ReportTemplate.EXECUTIVE_SUMMARY) -> bytes:
        """
        Export combined ARIMA + Optimization report
        
        Args:
            arima_data: ARIMA forecasting results
            optimization_data: Optimization results
            format: Export format
            template: Report template
            
        Returns:
            Exported combined report
        """
        logger.info(f"Exporting combined report in {format} format")
        
        combined_data = {
            'arima_results': arima_data,
            'optimization_results': optimization_data,
            'report_metadata': {
                'generated_at': datetime.now().isoformat(),
                'template': template,
                'version': '2.0'
            }
        }
        
        if format == ExportFormat.EXCEL:
            return self._export_combined_excel(combined_data, template)
        elif format == ExportFormat.PDF:
            return self._export_combined_pdf(combined_data, template)
        elif format == ExportFormat.ZIP:
            return self._export_combined_zip(combined_data)
        else:
            raise ValueError(f"Unsupported format for combined export: {format}")
    
    def export_charts(self, charts: List[go.Figure], format: str = ExportFormat.PNG,
                     width: int = 1200, height: int = 800) -> Union[bytes, List[bytes]]:
        """
        Export Plotly charts as images
        
        Args:
            charts: List of Plotly figures
            format: Image format (PNG, SVG, HTML)
            width: Image width
            height: Image height
            
        Returns:
            Exported chart data
        """
        logger.info(f"Exporting {len(charts)} charts in {format} format")
        
        exported_charts = []
        
        for i, chart in enumerate(charts):
            try:
                if format == ExportFormat.PNG:
                    img_bytes = pio.to_image(chart, format='png', width=width, height=height)
                    exported_charts.append(img_bytes)
                elif format == ExportFormat.SVG:
                    svg_str = pio.to_image(chart, format='svg', width=width, height=height)
                    exported_charts.append(svg_str)
                elif format == ExportFormat.HTML:
                    html_str = pio.to_html(chart, include_plotlyjs=True)
                    exported_charts.append(html_str.encode('utf-8'))
                else:
                    raise ValueError(f"Unsupported chart format: {format}")
                    
            except Exception as e:
                logger.error(f"Failed to export chart {i}: {str(e)}")
                continue
        
        if len(exported_charts) == 1:
            return exported_charts[0]
        else:
            return exported_charts
    
    def _export_arima_excel(self, arima_data: Dict[str, Any], template: str) -> bytes:
        """Export ARIMA results to Excel"""
        buffer = BytesIO()
        
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            # Summary sheet
            summary_data = self._create_arima_summary(arima_data)
            summary_df = pd.DataFrame([summary_data])
            summary_df.to_excel(writer, sheet_name='Executive Summary', index=False)
            
            # Forecast data for each location
            for location_id, location_data in arima_data.items():
                if isinstance(location_data, dict) and 'forecast' in location_data:
                    # Create forecast DataFrame
                    forecast_df = pd.DataFrame({
                        'Day': range(1, len(location_data['forecast']) + 1),
                        'Forecast (L)': location_data['forecast'],
                        'Upper CI': location_data.get('confidence_intervals', {}).get('upper', []),
                        'Lower CI': location_data.get('confidence_intervals', {}).get('lower', [])
                    })
                    
                    # Clean sheet name
                    sheet_name = f"Forecast_{location_id.replace('_', ' ')[:25]}"
                    forecast_df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # Apply styling
                    self._apply_excel_styling(writer, sheet_name, forecast_df)
            
            # Metrics comparison sheet
            metrics_data = self._extract_arima_metrics(arima_data)
            if metrics_data:
                metrics_df = pd.DataFrame(metrics_data)
                metrics_df.to_excel(writer, sheet_name='Model Metrics', index=False)
                self._apply_excel_styling(writer, 'Model Metrics', metrics_df)
        
        buffer.seek(0)
        return buffer.getvalue()
    
    def _export_optimization_excel(self, optimization_data: Dict[str, Any], template: str) -> bytes:
        """Export optimization results to Excel"""
        buffer = BytesIO()
        
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            # Executive summary
            summary_data = self._create_optimization_summary(optimization_data)
            summary_df = pd.DataFrame([summary_data])
            summary_df.to_excel(writer, sheet_name='Executive Summary', index=False)
            
            # Scenario comparison (if multiple scenarios)
            if 'scenario_results' in optimization_data:
                comparison_data = []
                for scenario_id, result in optimization_data['scenario_results'].items():
                    if 'optimization_summary' in result:
                        scenario_summary = {
                            'Scenario': result.get('scenario_info', {}).get('scenario_name', scenario_id),
                            'Fitness Score': result['optimization_summary']['optimization_fitness'],
                            'Total Cost (IDR)': result['route_optimization']['total_cost'],
                            'Total Distance (km)': result['route_optimization']['total_distance'],
                            'Service Level (%)': result['optimization_summary']['service_level_achieved'] * 100,
                            'Vehicles Used': result['route_optimization']['total_locations'],
                            'Optimization Time (s)': result.get('scenario_info', {}).get('optimization_time_seconds', 0)
                        }
                        comparison_data.append(scenario_summary)
                
                if comparison_data:
                    comparison_df = pd.DataFrame(comparison_data)
                    comparison_df.to_excel(writer, sheet_name='Scenario Comparison', index=False)
                    self._apply_excel_styling(writer, 'Scenario Comparison', comparison_df)
            
            # Route details
            self._add_route_details_to_excel(writer, optimization_data)
            
            # Location analysis
            self._add_location_analysis_to_excel(writer, optimization_data)
            
            # Convergence data
            self._add_convergence_data_to_excel(writer, optimization_data)
        
        buffer.seek(0)
        return buffer.getvalue()
    
    def _export_combined_excel(self, combined_data: Dict[str, Any], template: str) -> bytes:
        """Export combined report to Excel"""
        buffer = BytesIO()
        
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            # Executive Dashboard
            exec_data = self._create_executive_dashboard_data(combined_data)
            exec_df = pd.DataFrame([exec_data])
            exec_df.to_excel(writer, sheet_name='Executive Dashboard', index=False)
            
            # ARIMA Summary
            arima_summary = self._create_arima_summary(combined_data['arima_results'])
            arima_df = pd.DataFrame([arima_summary])
            arima_df.to_excel(writer, sheet_name='Forecast Summary', index=False)
            
            # Optimization Summary
            opt_summary = self._create_optimization_summary(combined_data['optimization_results'])
            opt_df = pd.DataFrame([opt_summary])
            opt_df.to_excel(writer, sheet_name='Optimization Summary', index=False)
            
            # Detailed forecasts
            self._add_forecast_details_to_excel(writer, combined_data['arima_results'])
            
            # Detailed routes
            self._add_route_details_to_excel(writer, combined_data['optimization_results'])
            
            # Combined insights
            insights_data = self._generate_combined_insights(combined_data)
            insights_df = pd.DataFrame(insights_data)
            insights_df.to_excel(writer, sheet_name='Key Insights', index=False)
            
            # Apply professional styling to all sheets
            for sheet_name in writer.sheets:
                if sheet_name in writer.sheets:
                    self._apply_professional_styling(writer, sheet_name)
        
        buffer.seek(0)
        return buffer.getvalue()
    
    def _export_arima_pdf(self, arima_data: Dict[str, Any], template: str) -> bytes:
        """Export ARIMA results to PDF"""
        if not PDF_AVAILABLE:
            raise RuntimeError("PDF export not available - install reportlab")
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=20,
            textColor=colors.HexColor(self.config['color_scheme']['primary']),
            alignment=1,  # Center
            spaceAfter=20
        )
        
        story.append(Paragraph("BBM Demand Forecasting Report", title_style))
        story.append(Spacer(1, 20))
        
        # Executive Summary
        story.append(Paragraph("Executive Summary", styles['Heading2']))
        summary_data = self._create_arima_summary(arima_data)
        
        summary_text = f"""
        This report presents ARIMA forecasting results for BBM demand analysis.
        
        • Total Locations Analyzed: {summary_data.get('Total Locations', 'N/A')}
        • Forecast Period: {summary_data.get('Forecast Days', 'N/A')} days
        • Total Forecasted Demand: {summary_data.get('Total Forecasted Demand (L)', 'N/A'):,} liters
        • Average Model Accuracy: {summary_data.get('Average MAPE (%)', 'N/A')}%
        """
        
        story.append(Paragraph(summary_text, styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Location-wise forecasts
        story.append(Paragraph("Location Forecasts", styles['Heading2']))
        
        for location_id, location_data in arima_data.items():
            if isinstance(location_data, dict) and 'forecast' in location_data:
                story.append(Paragraph(f"Location: {location_id}", styles['Heading3']))
                
                forecast_summary = f"""
                • Weekly Total Demand: {sum(location_data['forecast']):,.0f} liters
                • Daily Average: {np.mean(location_data['forecast']):,.0f} liters
                • Peak Day Demand: {max(location_data['forecast']):,.0f} liters
                • Model MAPE: {location_data.get('metrics', {}).get('mape', 'N/A')}%
                """
                
                story.append(Paragraph(forecast_summary, styles['Normal']))
                story.append(Spacer(1, 10))
        
        # Generate recommendations
        story.append(Paragraph("Recommendations", styles['Heading2']))
        recommendations = self._generate_arima_recommendations(arima_data)
        
        for i, rec in enumerate(recommendations, 1):
            story.append(Paragraph(f"{i}. {rec}", styles['Normal']))
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()
    
    def _export_optimization_pdf(self, optimization_data: Dict[str, Any], template: str) -> bytes:
        """Export optimization results to PDF"""
        if not PDF_AVAILABLE:
            raise RuntimeError("PDF export not available - install reportlab")
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=20,
            textColor=colors.HexColor(self.config['color_scheme']['primary']),
            alignment=1,
            spaceAfter=20
        )
        
        story.append(Paragraph("MARLOFIR-P Optimization Report", title_style))
        story.append(Spacer(1, 20))
        
        # Executive Summary
        story.append(Paragraph("Executive Summary", styles['Heading2']))
        summary_data = self._create_optimization_summary(optimization_data)
        
        summary_text = f"""
        This report presents the results of multi-objective genetic algorithm optimization
        for BBM distribution routing and scheduling.
        
        • Total Scenarios Analyzed: {summary_data.get('Scenarios Executed', 'N/A')}
        • Best Fitness Score: {summary_data.get('Best Fitness Score', 'N/A')}
        • Optimal Total Cost: IDR {summary_data.get('Total Cost (IDR)', 0):,.0f}
        • Total Distance: {summary_data.get('Total Distance (km)', 'N/A')} km
        • Service Level Achieved: {summary_data.get('Service Level (%)', 'N/A')}%
        • Vehicles Required: {summary_data.get('Vehicles Used', 'N/A')}
        """
        
        story.append(Paragraph(summary_text, styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Scenario comparison (if available)
        if 'scenario_results' in optimization_data:
            story.append(Paragraph("Scenario Analysis", styles['Heading2']))
            
            for scenario_id, result in optimization_data['scenario_results'].items():
                scenario_name = result.get('scenario_info', {}).get('scenario_name', scenario_id)
                story.append(Paragraph(f"Scenario: {scenario_name}", styles['Heading3']))
                
                if 'optimization_summary' in result:
                    opt_summary = result['optimization_summary']
                    route_opt = result['route_optimization']
                    
                    scenario_text = f"""
                    • Fitness Score: {opt_summary['optimization_fitness']:.2f}
                    • Total Cost: IDR {route_opt['total_cost']:,.0f}
                    • Total Distance: {route_opt['total_distance']:.1f} km
                    • Service Level: {opt_summary['service_level_achieved']:.1%}
                    """
                    
                    story.append(Paragraph(scenario_text, styles['Normal']))
                    story.append(Spacer(1, 10))
        
        # Recommendations
        story.append(Paragraph("Optimization Recommendations", styles['Heading2']))
        
        # Extract recommendations from any scenario
        recommendations = []
        for scenario_id, result in optimization_data.get('scenario_results', {}).items():
            if 'recommendations' in result:
                recommendations.extend(result['recommendations'])
                break
        
        if not recommendations:
            recommendations = ["Implement the optimized routes to achieve cost savings and improved service levels."]
        
        for i, rec in enumerate(recommendations, 1):
            story.append(Paragraph(f"{i}. {rec}", styles['Normal']))
        
        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()
    
    def _export_combined_zip(self, combined_data: Dict[str, Any]) -> bytes:
        """Export combined data as ZIP archive"""
        buffer = BytesIO()
        
        with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            # Add Excel report
            excel_data = self._export_combined_excel(combined_data, ReportTemplate.EXECUTIVE_SUMMARY)
            zip_file.writestr("BBM_Complete_Analysis_Report.xlsx", excel_data)
            
            # Add JSON data
            json_data = json.dumps(combined_data, indent=2, default=str)
            zip_file.writestr("BBM_Complete_Data.json", json_data)
            
            # Add ARIMA CSV
            arima_csv = self._create_arima_csv(combined_data['arima_results'])
            zip_file.writestr("ARIMA_Forecasts.csv", arima_csv)
            
            # Add optimization CSV
            opt_csv = self._create_optimization_csv(combined_data['optimization_results'])
            zip_file.writestr("Optimization_Results.csv", opt_csv)
            
            # Add README
            readme_content = self._create_readme_content()
            zip_file.writestr("README.txt", readme_content)
        
        buffer.seek(0)
        return buffer.getvalue()
    
    def _apply_excel_styling(self, writer: pd.ExcelWriter, sheet_name: str, df: pd.DataFrame):
        """Apply basic Excel styling"""
        try:
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]
            
            # Header styling
            header_fill = PatternFill(start_color=self.config['excel_style']['header_fill'],
                                    end_color=self.config['excel_style']['header_fill'],
                                    fill_type='solid')
            header_font = Font(name=self.config['excel_style']['header_font'], bold=True)
            
            # Apply header styling
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        except Exception as e:
            logger.warning(f"Failed to apply Excel styling: {str(e)}")
    
    def _apply_professional_styling(self, writer: pd.ExcelWriter, sheet_name: str):
        """Apply professional styling to Excel sheet"""
        try:
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]
            
            # Define styles
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
            data_font = Font(name='Calibri', size=11)
            border = Border(
                left=Side(border_style='thin'),
                right=Side(border_style='thin'),
                top=Side(border_style='thin'),
                bottom=Side(border_style='thin')
            )
            
            # Apply to all cells
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.font = data_font
                    cell.border = border
                    cell.alignment = Alignment(vertical='center')
            
            # Apply header styling
            if worksheet.max_row > 0:
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Auto-adjust columns
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        except Exception as e:
            logger.warning(f"Failed to apply professional styling: {str(e)}")
    
    def _create_arima_summary(self, arima_data: Dict[str, Any]) -> Dict[str, Any]:
        """Create ARIMA summary statistics"""
        total_locations = len(arima_data)
        total_forecast_demand = 0
        total_forecast_days = 0
        mape_values = []
        
        for location_id, location_data in arima_data.items():
            if isinstance(location_data, dict) and 'forecast' in location_data:
                total_forecast_demand += sum(location_data['forecast'])
                total_forecast_days = len(location_data['forecast'])
                
                if 'metrics' in location_data and 'mape' in location_data['metrics']:
                    mape_values.append(location_data['metrics']['mape'])
        
        return {
            'Total Locations': total_locations,
            'Forecast Days': total_forecast_days,
            'Total Forecasted Demand (L)': f"{total_forecast_demand:,.0f}",
            'Average Daily Demand (L)': f"{total_forecast_demand / (total_forecast_days * total_locations):,.0f}" if total_forecast_days > 0 else "N/A",
            'Average MAPE (%)': f"{np.mean(mape_values):.2f}" if mape_values else "N/A",
            'Report Generated': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
    
    def _create_optimization_summary(self, optimization_data: Dict[str, Any]) -> Dict[str, Any]:
        """Create optimization summary statistics"""
        if 'scenario_results' in optimization_data:
            # Multi-scenario results
            best_scenario = None
            best_fitness = 0
            
            for scenario_id, result in optimization_data['scenario_results'].items():
                if 'optimization_summary' in result:
                    fitness = result['optimization_summary']['optimization_fitness']
                    if fitness > best_fitness:
                        best_fitness = fitness
                        best_scenario = result
            
            if best_scenario:
                opt_summary = best_scenario['optimization_summary']
                route_opt = best_scenario['route_optimization']
                
                return {
                    'Scenarios Executed': len(optimization_data['scenario_results']),
                    'Best Fitness Score': f"{best_fitness:.2f}",
                    'Total Cost (IDR)': f"{route_opt['total_cost']:,.0f}",
                    'Total Distance (km)': f"{route_opt['total_distance']:.1f}",
                    'Service Level (%)': f"{opt_summary['service_level_achieved'] * 100:.1f}",
                    'Vehicles Used': route_opt.get('total_locations', 'N/A'),
                    'Demand Served (L)': f"{opt_summary['total_demand_served']:,.0f}",
                    'Report Generated': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                }
        
        # Single scenario result
        if 'optimization_summary' in optimization_data:
            opt_summary = optimization_data['optimization_summary']
            route_opt = optimization_data['route_optimization']
            
            return {
                'Scenarios Executed': 1,
                'Best Fitness Score': f"{opt_summary['optimization_fitness']:.2f}",
                'Total Cost (IDR)': f"{route_opt['total_cost']:,.0f}",
                'Total Distance (km)': f"{route_opt['total_distance']:.1f}",
                'Service Level (%)': f"{opt_summary['service_level_achieved'] * 100:.1f}",
                'Vehicles Used': route_opt.get('total_locations', 'N/A'),
                'Demand Served (L)': f"{opt_summary['total_demand_served']:,.0f}",
                'Report Generated': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
        
        return {'Error': 'Invalid optimization data format'}
    
    def _create_executive_dashboard_data(self, combined_data: Dict[str, Any]) -> Dict[str, Any]:
        """Create executive dashboard summary"""
        arima_summary = self._create_arima_summary(combined_data['arima_results'])
        opt_summary = self._create_optimization_summary(combined_data['optimization_results'])
        
        return {
            'Analysis Type': 'Combined ARIMA Forecasting + MARLOFIR-P Optimization',
            'Locations Analyzed': arima_summary.get('Total Locations', 0),
            'Forecast Period (days)': arima_summary.get('Forecast Days', 0),
            'Total Demand Forecasted (L)': arima_summary.get('Total Forecasted Demand (L)', '0'),
            'Forecast Accuracy (MAPE %)': arima_summary.get('Average MAPE (%)', 'N/A'),
            'Optimization Scenarios': opt_summary.get('Scenarios Executed', 0),
            'Best Solution Fitness': opt_summary.get('Best Fitness Score', '0'),
            'Optimized Total Cost (IDR)': opt_summary.get('Total Cost (IDR)', '0'),
            'Optimized Distance (km)': opt_summary.get('Total Distance (km)', '0'),
            'Service Level Achieved (%)': opt_summary.get('Service Level (%)', '0'),
            'Vehicles Required': opt_summary.get('Vehicles Used', 0),
            'Cost Savings Potential': self._calculate_cost_savings(combined_data),
            'Report Generated': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'System Version': 'MARLOFIR-P v2.0'
        }
    
    def _calculate_cost_savings(self, combined_data: Dict[str, Any]) -> str:
        """Calculate potential cost savings from optimization"""
        try:
            # Simple heuristic: compare optimized vs naive routing
            opt_data = combined_data['optimization_results']
            
            if 'scenario_results' in opt_data:
                # Find best scenario
                best_cost = float('inf')
                worst_cost = 0
                
                for scenario_id, result in opt_data['scenario_results'].items():
                    if 'route_optimization' in result:
                        cost = result['route_optimization']['total_cost']
                        best_cost = min(best_cost, cost)
                        worst_cost = max(worst_cost, cost)
                
                if worst_cost > best_cost:
                    savings_pct = ((worst_cost - best_cost) / worst_cost) * 100
                    return f"{savings_pct:.1f}% vs worst scenario"
            
            return "15-25% (estimated)"  # Default estimate
        
        except:
            return "Not calculated"
    
    def _extract_arima_metrics(self, arima_data: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Extract ARIMA metrics for comparison"""
        metrics_data = []
        
        for location_id, location_data in arima_data.items():
            if isinstance(location_data, dict) and 'metrics' in location_data:
                metrics = location_data['metrics']
                forecast = location_data.get('forecast', [])
                
                metrics_data.append({
                    'Location': location_id,
                    'MAPE (%)': metrics.get('mape', 'N/A'),
                    'RMSE': metrics.get('rmse', 'N/A'),
                    'AIC': metrics.get('aic', 'N/A'),
                    'Weekly Forecast (L)': sum(forecast) if forecast else 0,
                    'Daily Average (L)': np.mean(forecast) if forecast else 0,
                    'Forecast Std Dev': np.std(forecast) if forecast else 0
                })
        
        return metrics_data
    
    def _add_route_details_to_excel(self, writer: pd.ExcelWriter, optimization_data: Dict[str, Any]):
        """Add route details to Excel"""
        route_data = []
        
        # Extract route data from best scenario
        best_result = self._get_best_optimization_result(optimization_data)
        
        if best_result and 'location_analysis' in best_result:
            for location in best_result['location_analysis']:
                route_data.append({
                    'Location ID': location.get('location_id', ''),
                    'Location Name': location.get('location_name', ''),
                    'Forecasted Demand (L)': f"{location.get('forecasted_demand', 0):,.0f}",
                    'Current Stock (L)': f"{location.get('current_stock', 0):,.0f}",
                    'Urgency Score': f"{location.get('urgency_score', 0):.2f}",
                    'Priority Weight': f"{location.get('priority_weight', 0):.2f}"
                })
        
        if route_data:
            route_df = pd.DataFrame(route_data)
            route_df.to_excel(writer, sheet_name='Route Details', index=False)
            self._apply_excel_styling(writer, 'Route Details', route_df)
    
    def _add_location_analysis_to_excel(self, writer: pd.ExcelWriter, optimization_data: Dict[str, Any]):
        """Add location analysis to Excel"""
        best_result = self._get_best_optimization_result(optimization_data)
        
        if best_result and 'route_optimization' in best_result:
            route_opt = best_result['route_optimization']
            route_sequence = route_opt.get('route_sequence', [])
            
            if route_sequence:
                sequence_data = []
                for i, location in enumerate(route_sequence):
                    sequence_data.append({
                        'Stop Order': i + 1,
                        'Location': location,
                        'Location Type': 'Depot' if 'depot' in location.lower() else 'SPBU'
                    })
                
                sequence_df = pd.DataFrame(sequence_data)
                sequence_df.to_excel(writer, sheet_name='Route Sequence', index=False)
                self._apply_excel_styling(writer, 'Route Sequence', sequence_df)
    
    def _add_convergence_data_to_excel(self, writer: pd.ExcelWriter, optimization_data: Dict[str, Any]):
        """Add GA convergence data to Excel"""
        best_result = self._get_best_optimization_result(optimization_data)
        
        if best_result and 'convergence_history' in best_result:
            convergence_data = []
            for gen_data in best_result['convergence_history']:
                convergence_data.append({
                    'Generation': gen_data['generation'],
                    'Best Fitness': f"{gen_data['best_fitness']:.2f}",
                    'Average Fitness': f"{gen_data['avg_fitness']:.2f}",
                    'Population Diversity': f"{gen_data['diversity']:.3f}"
                })
            
            convergence_df = pd.DataFrame(convergence_data)
            convergence_df.to_excel(writer, sheet_name='Algorithm Convergence', index=False)
            self._apply_excel_styling(writer, 'Algorithm Convergence', convergence_df)
    
    def _add_forecast_details_to_excel(self, writer: pd.ExcelWriter, arima_data: Dict[str, Any]):
        """Add detailed forecast data to Excel"""
        all_forecasts = []
        
        for location_id, location_data in arima_data.items():
            if isinstance(location_data, dict) and 'forecast' in location_data:
                forecast = location_data['forecast']
                confidence_intervals = location_data.get('confidence_intervals', {})
                
                for day, value in enumerate(forecast, 1):
                    forecast_row = {
                        'Location': location_id,
                        'Day': day,
                        'Forecast (L)': f"{value:,.0f}",
                        'Upper CI (L)': f"{confidence_intervals.get('upper', [0] * len(forecast))[day-1]:,.0f}",
                        'Lower CI (L)': f"{confidence_intervals.get('lower', [0] * len(forecast))[day-1]:,.0f}"
                    }
                    all_forecasts.append(forecast_row)
        
        if all_forecasts:
            forecast_df = pd.DataFrame(all_forecasts)
            forecast_df.to_excel(writer, sheet_name='Detailed Forecasts', index=False)
            self._apply_excel_styling(writer, 'Detailed Forecasts', forecast_df)
    
    def _get_best_optimization_result(self, optimization_data: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        """Get best optimization result from data"""
        if 'scenario_results' in optimization_data:
            best_result = None
            best_fitness = 0
            
            for scenario_id, result in optimization_data['scenario_results'].items():
                if 'optimization_summary' in result:
                    fitness = result['optimization_summary']['optimization_fitness']
                    if fitness > best_fitness:
                        best_fitness = fitness
                        best_result = result
            
            return best_result
        
        elif 'optimization_summary' in optimization_data:
            return optimization_data
        
        return None
    
    def _generate_combined_insights(self, combined_data: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Generate key insights from combined analysis"""
        insights = []
        
        # Forecast insights
        arima_data = combined_data['arima_results']
        total_demand = 0
        peak_demand = 0
        
        for location_id, location_data in arima_data.items():
            if isinstance(location_data, dict) and 'forecast' in location_data:
                forecast = location_data['forecast']
                total_demand += sum(forecast)
                peak_demand = max(peak_demand, max(forecast))
        
        insights.append({
            'Category': 'Demand Forecasting',
            'Insight': f'Total weekly demand forecasted: {total_demand:,.0f} liters',
            'Impact': 'High',
            'Action Required': 'Ensure adequate inventory planning'
        })
        
        insights.append({
            'Category': 'Peak Demand',
            'Insight': f'Peak daily demand: {peak_demand:,.0f} liters',
            'Impact': 'Medium',
            'Action Required': 'Prepare for peak demand scenarios'
        })
        
        # Optimization insights
        opt_data = combined_data['optimization_results']
        best_result = self._get_best_optimization_result(opt_data)
        
        if best_result:
            cost = best_result.get('route_optimization', {}).get('total_cost', 0)
            service_level = best_result.get('optimization_summary', {}).get('service_level_achieved', 0)
            
            insights.append({
                'Category': 'Route Optimization',
                'Insight': f'Optimized total cost: IDR {cost:,.0f}',
                'Impact': 'High',
                'Action Required': 'Implement optimized routes'
            })
            
            insights.append({
                'Category': 'Service Level',
                'Insight': f'Achieved service level: {service_level:.1%}',
                'Impact': 'High' if service_level >= 0.9 else 'Medium',
                'Action Required': 'Monitor service delivery performance'
            })
        
        # Add general insights
        insights.append({
            'Category': 'System Integration',
            'Insight': 'ARIMA forecasting successfully integrated with GA optimization',
            'Impact': 'High',
            'Action Required': 'Continue integrated planning approach'
        })
        
        return insights
    
    def _generate_arima_recommendations(self, arima_data: Dict[str, Any]) -> List[str]:
        """Generate ARIMA-specific recommendations"""
        recommendations = []
        
        # Analyze forecast accuracy
        mape_values = []
        for location_data in arima_data.values():
            if isinstance(location_data, dict) and 'metrics' in location_data:
                mape = location_data['metrics'].get('mape')
                if mape is not None:
                    mape_values.append(mape)
        
        if mape_values:
            avg_mape = np.mean(mape_values)
            if avg_mape > 10:
                recommendations.append("Model accuracy is below target (MAPE > 10%). Consider model re-tuning or additional data sources.")
            elif avg_mape < 5:
                recommendations.append("Excellent model accuracy achieved (MAPE < 5%). Current models are reliable for planning.")
        
        # Analyze demand patterns
        total_demand = 0
        demand_variance = []
        
        for location_data in arima_data.values():
            if isinstance(location_data, dict) and 'forecast' in location_data:
                forecast = location_data['forecast']
                total_demand += sum(forecast)
                demand_variance.append(np.std(forecast))
        
        if demand_variance:
            avg_variance = np.mean(demand_variance)
            if avg_variance > 1000:
                recommendations.append("High demand variability detected. Consider implementing safety stock policies.")
        
        recommendations.append("Update forecasting models monthly to maintain accuracy.")
        recommendations.append("Monitor actual vs forecasted demand to validate model performance.")
        
        return recommendations
    
    def _create_arima_csv(self, arima_data: Dict[str, Any]) -> str:
        """Create CSV export of ARIMA data"""
        csv_data = []
        
        for location_id, location_data in arima_data.items():
            if isinstance(location_data, dict) and 'forecast' in location_data:
                forecast = location_data['forecast']
                confidence_intervals = location_data.get('confidence_intervals', {})
                metrics = location_data.get('metrics', {})
                
                for day, value in enumerate(forecast, 1):
                    csv_data.append({
                        'Location': location_id,
                        'Day': day,
                        'Forecast_L': value,
                        'Upper_CI_L': confidence_intervals.get('upper', [0] * len(forecast))[day-1],
                        'Lower_CI_L': confidence_intervals.get('lower', [0] * len(forecast))[day-1],
                        'MAPE': metrics.get('mape', ''),
                        'RMSE': metrics.get('rmse', ''),
                        'AIC': metrics.get('aic', '')
                    })
        
        df = pd.DataFrame(csv_data)
        return df.to_csv(index=False)
    
    def _create_optimization_csv(self, optimization_data: Dict[str, Any]) -> str:
        """Create CSV export of optimization data"""
        csv_data = []
        
        best_result = self._get_best_optimization_result(optimization_data)
        
        if best_result and 'location_analysis' in best_result:
            for location in best_result['location_analysis']:
                csv_data.append({
                    'Location_ID': location.get('location_id', ''),
                    'Location_Name': location.get('location_name', ''),
                    'Forecasted_Demand_L': location.get('forecasted_demand', 0),
                    'Current_Stock_L': location.get('current_stock', 0),
                    'Urgency_Score': location.get('urgency_score', 0),
                    'Priority_Weight': location.get('priority_weight', 0)
                })
        
        df = pd.DataFrame(csv_data)
        return df.to_csv(index=False)
    
    def _create_readme_content(self) -> str:
        """Create README content for ZIP export"""
        return """
BBM Distribution Analysis - Complete Export Package
===================================================

This package contains the complete analysis results from the BBM Dashboard and MARLOFIR-P optimization system.

Files Included:
- BBM_Complete_Analysis_Report.xlsx: Comprehensive Excel report with multiple sheets
- BBM_Complete_Data.json: Complete raw data in JSON format
- ARIMA_Forecasts.csv: ARIMA forecasting results in CSV format
- Optimization_Results.csv: Optimization analysis results in CSV format
- README.txt: This file

Excel Report Sheets:
- Executive Dashboard: High-level summary and KPIs
- Forecast Summary: ARIMA forecasting summary
- Optimization Summary: Route optimization summary
- Detailed Forecasts: Day-by-day forecast data
- Route Details: Optimized route information
- Key Insights: Strategic insights and recommendations

System Information:
- Analysis System: MARLOFIR-P v2.0
- Generated: """ + datetime.now().strftime('%Y-%m-%d %H:%M:%S') + """
- Components: ARIMA Forecasting + Genetic Algorithm Optimization

For questions or support, please contact the BBM Analytics team.
        """
    
    def _export_arima_json(self, arima_data: Dict[str, Any]) -> bytes:
        """Export ARIMA data as JSON"""
        export_data = {
            'export_metadata': {
                'export_type': 'arima_forecasting',
                'timestamp': datetime.now().isoformat(),
                'version': '2.0'
            },
            'summary': self._create_arima_summary(arima_data),
            'forecasting_results': arima_data,
            'metrics_comparison': self._extract_arima_metrics(arima_data)
        }
        
        return json.dumps(export_data, indent=2, default=str).encode('utf-8')
    
    def _export_optimization_json(self, optimization_data: Dict[str, Any]) -> bytes:
        """Export optimization data as JSON"""
        export_data = {
            'export_metadata': {
                'export_type': 'marlofir_optimization',
                'timestamp': datetime.now().isoformat(),
                'version': '2.0'
            },
            'summary': self._create_optimization_summary(optimization_data),
            'optimization_results': optimization_data
        }
        
        return json.dumps(export_data, indent=2, default=str).encode('utf-8')
    
    def _export_arima_html(self, arima_data: Dict[str, Any], template: str) -> bytes:
        """Export ARIMA results as HTML report"""
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>BBM ARIMA Forecasting Report</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 40px; }}
                .header {{ background-color: {self.config['color_scheme']['primary']}; color: white; padding: 20px; text-align: center; }}
                .summary {{ background-color: #f5f5f5; padding: 15px; margin: 20px 0; }}
                .location {{ border: 1px solid #ddd; margin: 10px 0; padding: 15px; }}
                .metric {{ display: inline-block; margin: 10px; padding: 10px; background-color: #e9ecef; }}
                table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
                th, td {{ border: 1px solid #ddd; padding: 12px; text-align: left; }}
                th {{ background-color: {self.config['color_scheme']['primary']}; color: white; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>BBM Demand Forecasting Report</h1>
                <p>Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
            </div>
            
            <div class="summary">
                <h2>Executive Summary</h2>
        """
        
        # Add summary data
        summary = self._create_arima_summary(arima_data)
        for key, value in summary.items():
            html_content += f"<div class='metric'><strong>{key}:</strong> {value}</div>"
        
        html_content += "</div>"
        
        # Add location details
        html_content += "<h2>Location Forecasts</h2>"
        
        for location_id, location_data in arima_data.items():
            if isinstance(location_data, dict) and 'forecast' in location_data:
                html_content += f"""
                <div class="location">
                    <h3>{location_id}</h3>
                    <p><strong>Weekly Total:</strong> {sum(location_data['forecast']):,.0f} liters</p>
                    <p><strong>Daily Average:</strong> {np.mean(location_data['forecast']):,.0f} liters</p>
                """
                
                if 'metrics' in location_data:
                    metrics = location_data['metrics']
                    html_content += f"<p><strong>Model MAPE:</strong> {metrics.get('mape', 'N/A')}%</p>"
                
                html_content += "</div>"
        
        html_content += """
            </body>
        </html>
        """
        
        return html_content.encode('utf-8')
    
    def get_export_statistics(self) -> Dict[str, Any]:
        """Get export performance statistics"""
        return {
            'performance_stats': self.export_stats,
            'supported_formats': [
                ExportFormat.EXCEL,
                ExportFormat.PDF,
                ExportFormat.JSON,
                ExportFormat.CSV,
                ExportFormat.HTML,
                ExportFormat.PNG,
                ExportFormat.SVG,
                ExportFormat.ZIP
            ],
            'available_templates': [
                ReportTemplate.EXECUTIVE_SUMMARY,
                ReportTemplate.TECHNICAL_DETAILS,
                ReportTemplate.COMPARISON_REPORT,
                ReportTemplate.OPTIMIZATION_ANALYSIS,
                ReportTemplate.FORECAST_REPORT
            ],
            'configuration': self.config
        }

# Utility functions for common export operations
def create_download_link(data: bytes, filename: str, mime_type: str) -> str:
    """Create base64 download link for Streamlit"""
    b64_data = base64.b64encode(data).decode()
    return f"data:{mime_type};base64,{b64_data}"

def get_mime_type(format: str) -> str:
    """Get MIME type for export format"""
    mime_types = {
        ExportFormat.EXCEL: "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ExportFormat.PDF: "application/pdf",
        ExportFormat.JSON: "application/json",
        ExportFormat.CSV: "text/csv",
        ExportFormat.HTML: "text/html",
        ExportFormat.PNG: "image/png",
        ExportFormat.SVG: "image/svg+xml",
        ExportFormat.ZIP: "application/zip"
    }
    return mime_types.get(format, "application/octet-stream")

# Example usage and testing
if __name__ == "__main__":
    # Initialize export utilities
    export_utils = ExportUtilities()
    
    print("Export Utilities Test Results:")
    print("=" * 50)
    
    # Test configuration
    print("\n1. Configuration Test:")
    stats = export_utils.get_export_statistics()
    print(f"   Supported formats: {len(stats['supported_formats'])}")
    print(f"   Available templates: {len(stats['available_templates'])}")
    print(f"   Color scheme: {stats['configuration']['color_scheme']['primary']}")
    
    # Create sample ARIMA data
    sample_arima_data = {
        'SPBU_001': {
            'forecast': [5000, 5200, 4800, 5100, 5300, 4900, 5050],
            'confidence_intervals': {
                'upper': [5500, 5700, 5300, 5600, 5800, 5400, 5550],
                'lower': [4500, 4700, 4300, 4600, 4800, 4400, 4550]
            },
            'metrics': {'mape': 4.2, 'rmse': 245.3, 'aic': 185.7}
        },
        'SPBU_002': {
            'forecast': [3500, 3600, 3400, 3550, 3650, 3450, 3575],
            'confidence_intervals': {
                'upper': [3850, 3960, 3740, 3905, 4015, 3795, 3932],
                'lower': [3150, 3240, 3060, 3195, 3285, 3105, 3218]
            },
            'metrics': {'mape': 5.8, 'rmse': 198.4, 'aic': 167.2}
        }
    }
    
    # Test ARIMA summary creation
    print("\n2. ARIMA Summary Test:")
    arima_summary = export_utils._create_arima_summary(sample_arima_data)
    for key, value in arima_summary.items():
        print(f"   {key}: {value}")
    
    # Test export formats
    print("\n3. Export Format Tests:")
    
    try:
        # Test JSON export
        json_data = export_utils.export_arima_results(sample_arima_data, ExportFormat.JSON)
        json_obj = json.loads(json_data.decode('utf-8'))
        print(f"   ✅ JSON export: {len(json_obj)} top-level keys")
        
        # Test Excel export
        excel_data = export_utils.export_arima_results(sample_arima_data, ExportFormat.EXCEL)
        print(f"   ✅ Excel export: {len(excel_data)} bytes")
        
        # Test HTML export
        html_data = export_utils.export_arima_results(sample_arima_data, ExportFormat.HTML)
        print(f"   ✅ HTML export: {len(html_data)} bytes")
        
        # Test CSV creation
        csv_data = export_utils._create_arima_csv(sample_arima_data)
        csv_lines = len(csv_data.split('\n'))
        print(f"   ✅ CSV export: {csv_lines} rows")
        
    except Exception as e:
        print(f"   ❌ Export test failed: {str(e)}")
    
    # Test PDF export (if available)
    print("\n4. PDF Export Test:")
    if PDF_AVAILABLE:
        try:
            pdf_data = export_utils.export_arima_results(sample_arima_data, ExportFormat.PDF)
            print(f"   ✅ PDF export: {len(pdf_data)} bytes")
        except Exception as e:
            print(f"   ❌ PDF export failed: {str(e)}")
    else:
        print("   ⚠️ PDF export not available (install reportlab)")
    
    # Test utility functions
    print("\n5. Utility Functions Test:")
    test_data = b"test data"
    download_link = create_download_link(test_data, "test.txt", "text/plain")
    print(f"   Download link created: {len(download_link)} characters")
    
    mime_type = get_mime_type(ExportFormat.EXCEL)
    print(f"   Excel MIME type: {mime_type}")
    
    print("\n" + "=" * 50)
    print("Export utilities test completed! 📊")